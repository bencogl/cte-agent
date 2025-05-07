from openpyxl import load_workbook
from datetime import datetime as D

def to_num(value):
    try:
        return float(value)
    except ValueError:
        return value

def extract_from_excel(filepath):
    wb = load_workbook(filepath, data_only=True)
    data = {}
    for ws in wb.worksheets:
        if ws['A1'].value == 'Codice prodotto':
            for row in ws.iter_rows(min_row=2, values_only=True):
                codice_listino = row[3] if row[3] else ""
                
                # Assicurati di avere un Codice Listino
                if not codice_listino:
                    continue
                
                listino_data = {
                    "filename": filepath,
                    "Codice Listino": codice_listino,
                    "Prodotto": row[1] if row[1] else "",
                    "Data fine vendibilità": row[10].strftime('%d/%m/%Y') if isinstance(row[10], D) else ""
                }

                # Conversione delle colonne numeriche
                for i, cell in enumerate(row):
                    if cell is not None and isinstance(cell, (int, float, D)):
                        listino_data[f"Colonna_{i}"] = to_num(str(cell))
                
                # Usa il Codice Listino come chiave
                data[codice_listino] = listino_data

    return data
