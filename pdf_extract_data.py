from glob import glob
from pprint import pprint

import ctevalidator

def text():
    for i, page in enumerate(data):
        len_page = len(page)
        for j, line in enumerate(page):
            print([i, f'{j}|{j - len_page}', line])

if __name__  == '__main__':
    
    pdf_files = ['INPUT\\NEXTENERGYSunlightDUAL_LUCE_100125.pdf']

    for filename in pdf_files:
        data = ctevalidator.extractFromPdf(filename)
        #pprint(data)
        #input()
    
    text()
    '''
    from openpyxl import load_workbook
    wb = load_workbook(filename = 'Tracciato caricatore_CorporateVendita_Settembre24.xlsx', data_only=True)
    print(wb)
    for ws in wb.worksheets:
        
        for value in ws.iter_rows(
            values_only=True):
            print(value)
    '''
    '''

        cell = ws['F3']
        print(cell.value)
    '''
