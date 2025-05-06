from openpyxl import load_workbook
from decimal import Decimal as D
from itertools import groupby
from collections import defaultdict

def parse_excel(file_path):
    wb = load_workbook(file_path, data_only=True)
    wss = [ws for ws in wb.worksheets if ws['A1'].value == 'Codice prodotto']
    all_data = []
    for ws in wss:
        for row in ws.iter_rows(min_row=2, values_only=True):
            all_data.append((file_path, ) + row)
    return all_data

def group_by_listino(rows):
    grouped = defaultdict(list)
    rows.sort(key=lambda x: (x[0], x[3]))
    for k, g in groupby(rows, key=lambda x: (x[0], x[3])):
        if k[1]:
            grouped[k].extend(list(g))
    return grouped

def interpret_excel_group(key, rows):
    try:
        price_name = key[1]
        other = {'xls': key[0]}
        p = {}
        for i, x in enumerate(rows):
            if i == 0:
                p['Codice Listino'] = x[3]
                p['Prodotto'] = x[1]
                if x[10]:
                    p['Data fine vendibilità'] = x[10].strftime('%d/%m/%Y')
            if x[5] is None or x[5].strip() == '':
                p[x[2]] = D(str(x[6]))
            else:
                p[f'{x[2]}|{x[5]}'] = D(str(x[6]))
            if x[7] is not None:
                p['Percentuale prezzo fisso'] = D(str(x[7]))
        return price_name, p, other
    except Exception as e:
        raise ValueError(f"Errore nel parsing Excel: {e}")