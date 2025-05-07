import csv
import datetime
import sys

from collections import Counter
from decimal import Decimal as D
from itertools import groupby
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from rules import *

input_dir = Path().absolute() / 'INPUT'

def toNum(s):
    return D(s.replace('.', '').replace(',', '.'))

def dictsDiff(a, b):
    return {x for x in a.keys() | b.keys() if a.get(x, None) != b.get(x, None)}

class Log:
    def __init__(self, filename):
        self.filename = filename
        self.write('File pdf', 'File xls', 'Listino', 'Tipo', 'Descrizione')

    def write(self, pdf_file=None, xls_file=None, listino=None, tipo=None, desc=None):
        with open(self.filename, 'a', newline='') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerow((pdf_file, xls_file, listino, tipo, desc))

def extractFromPdf(filename):
    reader = PdfReader(filename)
    return [p.extract_text().splitlines() for p in reader.pages]



if __name__  == '__main__':
    now = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    log2 = Log(f'log_{now}.csv')
    log2.write(tipo='Inizio')


    '''
    FASE 1: Analisi CTE in pdf
    '''

    print('lettura delle CTE', end='', flush=True)

    g001 = []
    pdf_files = input_dir.rglob('**/*.pdf')
    for pdf_file in pdf_files:

        print('.', end='', flush=True)
        log2.write(pdf_file=pdf_file.name,
                   tipo='Elaborazione',
                   desc='File in elaborazione')
        try:
            fulltext = extractFromPdf(pdf_file)
        except:
            raise
            log2.write(pdf_file=pdf_file.name,
                       tipo='Errore',
                       desc='Estrazione testo non riuscita')
            continue
        template = getInterpreter(pdf_file.name, fulltext)
        if template is None:
            log2.write(pdf_file=pdf_file.name, tipo='Errore', \
                       desc='Template non trovato')
            continue        
        try:
            content = template(fulltext)
        except:
            log2.write(pdf_file=pdf_file.name,
                       tipo='Errore',
                       desc='Estrazione dati listino non riuscita')
            continue

        g001 = g001 + [[pdf_file.name, template.__name__, c] for c in content]

    

    #formattazione corretta
    def f002(filename, template, data):
        pricelist = data.get('Codice Listino', None)
        pricing = {x: y for x, y in data.items() if x not in ['Nome Listino', 'Codice Offerta']}
        other = {'pdf': filename, 'template': template, \
                 'Nome Listino': data.get('Nome Listino', None), \
                 'Codice Offerta': data.get('Codice Offerta', None)}
        return pricelist, pricing, other 
    g002 = [f002(*x) for x in g001]

    print(' Esclusione duplicati', end='', flush=True)
    
    frequency = Counter(l for l, p, o in g002)

    pdf_data = {}
    for l, p, o in g002:
        if frequency[l] > 1:
            log2.write(pdf_file=o['pdf'],
                       tipo='Errore',
                       listino=l,
                       desc='Listino duplicato, tolto dalla lista')
            print('.', end='', flush=True)
        else:
            pdf_data[l] = (p, o)         
    del g002

    print('', end='\n', flush=True)



    '''
    FASE 2: Tracciati listini in xls 
    '''

    print('\nlettura dei tracciati di caricamento', end='', flush=True)

    full_data = []

    xls_files = input_dir.rglob('*.xlsx')
    for xls_file in xls_files:
        print('.', end='', flush=True)
        log2.write(xls_file=xls_file.name,
                   tipo='Elaborazione',
                   desc='File in elaborazione')
        try:
            wb = load_workbook(xls_file, data_only=True)
            wss = [x for x in wb.worksheets if x['A1'].value == 'Codice prodotto']
            datasheets = [list(ws.iter_rows(min_row=2, values_only=True)) for ws in wss]
            full_data = full_data + [(xls_file.name, ) + y for x in datasheets for y in x]
        except PermissionError:
            log2.write(xls_file=xls_file.name,
                tipo='Errore',
               desc='Impossibile utilizzare il file - è aperto in excel?')
            continue
        except:
            log2.write(xls_file=xls_file.name,
                tipo='Errore',
               desc='Errore sconosciuto')
            continue

    
    g005 = groupby(full_data, lambda x: (x[0], x[3])) #filename e listino
    g006 = ((k, v) for k, v in g005 if k[1] is not None and k[1] != '')

    def f007(k, v):
        try:
            price_name = k[1]
            other = {'xls': k[0]}
            p = {}
            for i, x in enumerate(v):
                if i == 0:
                    p['Codice Listino'] = x[3]
                    p['Prodotto'] = x[1]
                    p['Data fine vendibilità'] = '' if x[10] is None else x[10].strftime('%d/%m/%Y')                
                if x[5] is None or x[5].strip() == '':
                    p[x[2]] = D(str(x[6]))
                else:
                    p[x[2] + '|' + x[5]] = D(str(x[6]))            
                if x[7] is not None:
                    p['Percentuale prezzo fisso'] = D(str(x[7]))
            return price_name, p, other
        except:
            print((k, list(v)), flush=True)

    g007 = [f007(k, v) for k, v in g006]

    print(' Esclusione duplicati', end='', flush=True)
    frequency = Counter(l for l, p, o in g007)
    xls_data = {}
    for l, p, o in g007:
        if frequency[l] > 1:
            log2.write(xls_file=o['xls'],
                       tipo='Errore',
                       listino=l,
                       desc='Listino duplicato, tolto dalla lista')
            print('.', end='', flush=True)
        else:
            xls_data[l] = (p, o)         
    del g007

    print('', end='\n', flush=True)



    '''
    FASE 3: Analisi differenze 
    '''

    print('\nAnalisi coerenza listini', flush=True)

    listini_identici = set()
    for x in pdf_data.keys() | xls_data.keys():
        if x not in pdf_data.keys():
            log2.write(None, xls_data[x][1]['xls'], x, 'Errore', 'Non trovato pdf')
            print(f'KO :( {x}', flush=True)
            continue
        if x not in xls_data.keys():
            log2.write(pdf_data[x][1]['pdf'], None, x, 'Errore', 'Non trovato tracciato xls')
            print(f'KO :( {x}', flush=True)            
            continue
        diffs = dictsDiff(pdf_data[x][0], xls_data[x][0])
        if len(diffs) > 0:
            log2.write(pdf_data[x][1]['pdf'], xls_data[x][1]['xls'], x, 'Errore', f'Trovate differenze tra pdf e xls: {diffs}')
            print(f'KO :( {x}', flush=True)
        else:
            log2.write(pdf_data[x][1]['pdf'], xls_data[x][1]['xls'], x, 'Elaborazione', 'Coerenza verificata con successo')
            print(f'OK    {x}', flush=True)
            listini_identici.add(x)

    if len(listini_identici) == 0:
        log2.write(tipo='Errore', desc='Nessun abbinamento trovato!')
    else:
        log2.write(tipo='Elaborazione', desc='Salvataggio abbinamenti riusciti')
        with open(f'sintesi_{now}.csv', 'w', newline='') as f_9:
            writer = csv.writer(f_9, delimiter=';')
            header_1 = ['pdf', 'Codice Offerta', 'template', 'Nome Listino']
            header_2 = ['xls']
            header_3 = sorted(set(y for x in listini_identici for y in pdf_data[x][0].keys()))
            writer.writerow(header_1 + header_2 + header_3)
            for x in listini_identici:
                line = [pdf_data[x][1].get(c, '') for c in header_1] + \
                       [xls_data[x][1].get(c, '') for c in header_2] + \
                       [pdf_data[x][0].get(c, '') for c in header_3]                   
                writer.writerow(line)

    log2.write(tipo='Fine')

    input('\nPremi INVIO per uscire')
